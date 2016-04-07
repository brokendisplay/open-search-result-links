import openpyxl
import webbrowser
from openpyxl.styles import PatternFill

# figure out a way to generate the url name (same name, but appended
# date), so perhaps something like
# <generic_file_name>_<formatted_date>.xlsx
workbookName = 'elearning.xlsx'

workbook = openpyxl.load_workbook(workbookName)
worksheet = workbook.get_sheet_by_name('Sheet1')
sheetValue = 1  # initialize
# linkRow = 1
linkColumn = 5  # the column where the links are
maxRow = worksheet.max_row
maxColumn = worksheet.max_col

# initialize the response variables
privateResponse = 0
customerResponse = 0
# need to figure out a way to compare an empty str to None type

# define the types of fills that we need
# privateFill
privateFill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
# publicCustomerFIll
publicCustomerFill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
# LMS fill
LMSfill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
# we could make use of a while loop...

# while sheetValue is not None:  # while there is data in that cell
# 	sheetValue = worksheet.cell(row=linkRow, column=linkColumn).value
# 	webbrowser.open(sheetValue)  # open the link in Internet Explorer
# 	# print(sheetValue)
# 	linkRow += 1  # increment by 1
# print("All links have been opened.")  # finish

# or we could use the maxRow parameter in a for loop!

# if I want to open up all links at once, I'll have to throw in a halt
# if I make it interactive, I'll have to figure out how to apply a
# style (background color) to the row

for i in range(1, maxRow+1):
	sheetValue = worksheet.cell(row=i, column=linkColumn).value
	webbrowser.open(sheetValue)
	# prompt the user with two questions -- change the row
	# background depending on the user's response
	while privateResponse != "Y":
		privateResponse = input("Is this link private? [Y/N] ").upper()
		# check user's response
		if privateResponse != "Y":
			# color the row green (false positive)
			for j in range(1, maxColumn+1):
				cellRef = worksheet.cell(row=i, column=j)
				worksheet[cellRef] = privateFill
		elif privateResponse == "N":
			while customerResponse != "Y":
				customerResponse = input("Does this site belong to a customer? [Y/N] ")
			if customerResponse == "Y":
				# color the row orange (I will contact)
				for k in range(1, maxColumn+1):
					cellRef = worksheet.cell(row=i, column=k)
					worksheet[cellRef] = publicCustomerFill
			else:  # if an LMS
				# color the row yellow (vendor relations)
				for l in range(1, maxColumn+1):
					cellRef = worksheet.cell(row=i, column=l)
					worksheet[cellRef] = LMSfill
		else:
			print("Invalid value. Try again.")
			break

print("All links have been opened.")
